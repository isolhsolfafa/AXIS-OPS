"""
Material Parser — CSV/xlsx 파싱 + 인코딩 감지 + Q1 MFC 합침 + 7종 검증 (Sprint 66-BE-FOLLOWUP v3)

Codex 5라운드 검증 GREEN 정합:
  - Q1 MFC scope only (category=='MFC' 영역 영역 합침)
  - non-MFC 중복 → dedup 첫 등장 사용 (053a `dedup_material_master()` 패턴)
  - ATTRIBUTE_CONFLICT → 같은 item_code 영역 자재 정보 충돌 (053a `validate_source_keys()` 패턴, 첫 등장 유지 + 후속 reject)
  - INVALID_BOM_KEY → BOM row 영역만 (product_code != '' 영역 영역)
  - FIELD_TOO_LONG → 8 필드 (item_code 50 / item_name 200 / category 50 / spec_1 200 / spec_2 200 / unit 20 / customer 100 / model 100)
  - 파일 형식: csv + xlsx (.xls drop)
"""
import csv
import io
import logging
from typing import Dict, List, Tuple, Optional, Any

import chardet
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


# 053a generator CSV_COLUMN_MAP 재사용 (한글 헤더 11컬럼)
CSV_COLUMN_MAP = {
    '품번': 'product_code',
    '고객사': 'customer',
    '모델': 'model',
    '자재코드': 'item_code',
    '자재내역': 'item_name',
    '규격1': 'spec_1',
    '규격2': 'spec_2',
    '수량': 'quantity',
    '단위': 'unit',
    '생성일': '_ignored',
    '비고': 'description',
}

# 영문 헤더도 직접 허용 (역방향)
ENGLISH_HEADERS = set(CSV_COLUMN_MAP.values()) - {'_ignored'}

# FIELD_TOO_LONG 검증 영역 8 필드 (description 영역 053b TEXT — 검증 X)
FIELD_MAX_LENGTH = {
    'item_code': 50,
    'item_name': 200,
    'category': 50,
    'spec_1': 200,
    'spec_2': 200,
    'unit': 20,
    'customer': 100,
    'model': 100,
}


def detect_encoding(file_bytes: bytes) -> str:
    """인코딩 자동 감지 — chardet → UTF-8 → CP949 → EUC-KR fallback.

    Raises:
        ValueError('ENCODING_DETECTION_FAILED'): 모든 fallback 영역 실패 시
    """
    if not file_bytes:
        return 'utf-8'

    result = chardet.detect(file_bytes)
    encoding = (result.get('encoding') or '').lower()
    confidence = result.get('confidence') or 0

    # chardet 신뢰도 ≥ 0.7 — 그대로 사용 (실 decode 시도)
    if confidence >= 0.7 and encoding:
        try:
            file_bytes.decode(encoding)
            return encoding
        except (UnicodeDecodeError, LookupError):
            pass

    # fallback 순서
    for candidate in ('utf-8', 'cp949', 'euc-kr'):
        try:
            file_bytes.decode(candidate)
            return candidate
        except UnicodeDecodeError:
            continue

    raise ValueError('ENCODING_DETECTION_FAILED')


def _parse_csv(file_bytes: bytes) -> List[Dict[str, str]]:
    """CSV 파싱 — 인코딩 자동 감지 후 한글/영문 헤더 모두 허용."""
    encoding = detect_encoding(file_bytes)
    text = file_bytes.decode(encoding)
    reader = csv.DictReader(io.StringIO(text))
    return [{(k or '').strip(): (v or '').strip() for k, v in row.items()} for row in reader]


def _parse_xlsx(file_bytes: bytes) -> List[Dict[str, str]]:
    """xlsx 파싱 — openpyxl 첫 번째 시트만 read.

    .xls 영역 영역 호출 → openpyxl 영역 InvalidFileException 영역 → 상위 PARSE_ERROR raise.
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError('PARSE_ERROR') from e

    ws = wb.active
    if ws is None:
        return []

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    headers = [(str(h) if h is not None else '').strip() for h in header_row]
    result: List[Dict[str, str]] = []
    for raw_row in rows_iter:
        # 빈 row 영역 skip (모든 셀 None/'')
        if not any(v is not None and str(v).strip() for v in raw_row):
            continue
        mapped = {}
        for i, cell in enumerate(raw_row):
            if i >= len(headers):
                break
            mapped[headers[i]] = (str(cell) if cell is not None else '').strip()
        result.append(mapped)
    return result


def _map_korean_to_english(row: Dict[str, str]) -> Dict[str, str]:
    """한글/영문 헤더 영역 영문 컬럼 영역 매핑."""
    mapped: Dict[str, str] = {}
    for key, value in row.items():
        eng_key = CSV_COLUMN_MAP.get(key, key)  # 한글 → 영문, 영문 → 영문 (passthrough)
        if eng_key == '_ignored':
            continue
        mapped[eng_key] = value
    return mapped


def _validate_row(mapped: Dict[str, str], row_number: int) -> Tuple[Dict[str, Any], Optional[Dict[str, Any]]]:
    """각 row 검증 — 7종 reject reason 영역 분류 (DUPLICATE_ITEM_CODE 영역 사용 X).

    Returns: (mapped_with_category, reject_or_None)
    """
    # 1) MISSING_ITEM_CODE
    if not mapped.get('item_code'):
        return mapped, {'row_number': row_number, 'reason': 'MISSING_ITEM_CODE'}

    # 2) MISSING_ITEM_NAME
    if not mapped.get('item_name'):
        return mapped, {'row_number': row_number, 'reason': 'MISSING_ITEM_NAME'}

    # 3) INVALID_QUANTITY (NULL 허용)
    qty = (mapped.get('quantity') or '').strip()
    if qty and not qty.lstrip('-').isdigit():
        return mapped, {'row_number': row_number, 'reason': 'INVALID_QUANTITY'}

    # 4) category 자동 추출 (053a generator 패턴)
    item_name = mapped['item_name']
    if item_name.upper().startswith('MFC'):
        mapped['item_name'] = 'MFC'
        mapped['category'] = 'MFC'
    else:
        mapped['category'] = item_name  # ⚠️ 50자 초과 시 FIELD_TOO_LONG (다음 영역 검증)

    # 5) FIELD_TOO_LONG (8 필드 검증, description 영역 TEXT — 검증 X)
    for field, max_len in FIELD_MAX_LENGTH.items():
        value = (mapped.get(field) or '')
        if len(value) > max_len:
            return mapped, {
                'row_number': row_number,
                'reason': 'FIELD_TOO_LONG',
                'detail': f'{field}={len(value)} chars > {max_len}',
            }

    # 6) INVALID_BOM_KEY — BOM row 영역만 적용 (product_code != '')
    # material-only MFC rows (product_code='') 영역 허용 (053a 패턴 정합)
    product_code = (mapped.get('product_code') or '').strip()
    if product_code:
        # BOM row → customer/model 필수
        if not (mapped.get('customer') or '').strip() or not (mapped.get('model') or '').strip():
            return mapped, {
                'row_number': row_number,
                'reason': 'INVALID_BOM_KEY',
                'detail': 'product_code present but customer/model missing',
            }

    return mapped, None


def _merge_duplicate_mfc(rows: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """동일 item_code 영역 description 합침 (Q1 — MFC scope only).

    v3 정정 (Codex 라운드 2 M-2):
      - MFC-only (category == 'MFC') 영역 합침 — description 합쳐서 단일 row
      - non-MFC 중복 — 첫 등장 row 사용 (053a `dedup_material_master()` 패턴)
      - 자재 정보 (item_name/spec_*/unit) 충돌 → ATTRIBUTE_CONFLICT reject (첫 등장 유지 + 후속 reject)

    Returns: (merged_rows, attribute_conflict_rejects)
    """
    seen: Dict[str, Dict[str, Any]] = {}
    rejects: List[Dict[str, Any]] = []

    for row in rows:
        item_code = row['item_code']
        if item_code not in seen:
            seen[item_code] = row
            continue

        existing = seen[item_code]
        # 자재 정보 충돌 검증 (053a validate_source_keys 패턴)
        conflict_field = None
        for field in ('item_name', 'spec_1', 'spec_2', 'unit'):
            if (existing.get(field) or '') != (row.get(field) or ''):
                conflict_field = field
                break

        if conflict_field:
            rejects.append({
                'row_number': row.get('_row_number'),
                'reason': 'ATTRIBUTE_CONFLICT',
                'detail': f'item_code={item_code} field={conflict_field} conflict',
            })
            continue

        # 충돌 X — MFC 영역만 description 합침
        if existing.get('category') == 'MFC' and row.get('category') == 'MFC':
            existing_desc = (existing.get('description') or '').strip()
            new_desc = (row.get('description') or '').strip()
            merged_tokens = sorted({
                tok.strip() for tok in (existing_desc.split(',') + new_desc.split(',')) if tok.strip()
            })
            existing['description'] = ','.join(merged_tokens)
        # non-MFC 영역 첫 등장 row 사용 (skip — reject X)

    return list(seen.values()), rejects


def parse_upload_file(file) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """파일 파싱 — CSV / xlsx 자동 분기 + 검증 + Q1 MFC 합침.

    Returns: (parsed_rows, rejected_rows)

    Raises:
        ValueError: ENCODING_DETECTION_FAILED / INVALID_HEADER / PARSE_ERROR
    """
    filename = (file.filename or '').lower()
    file_bytes = file.read()

    if filename.endswith('.xlsx'):
        rows = _parse_xlsx(file_bytes)
    elif filename.endswith('.csv'):
        rows = _parse_csv(file_bytes)
    else:
        raise ValueError('PARSE_ERROR')  # .xls / 다른 확장자 영역 reject

    # 헤더 검증 — 한글/영문 모두 허용
    if rows:
        first = rows[0]
        keys = set(first.keys())
        if '자재코드' not in keys and 'item_code' not in keys:
            raise ValueError('INVALID_HEADER')

    parsed_rows: List[Dict[str, Any]] = []
    rejected_rows: List[Dict[str, Any]] = []

    for idx, row in enumerate(rows, start=2):  # row 1 = header
        mapped = _map_korean_to_english(row)
        mapped['_row_number'] = idx  # ATTRIBUTE_CONFLICT trail 영역
        validated, reject = _validate_row(mapped, idx)
        if reject:
            rejected_rows.append(reject)
            continue
        parsed_rows.append(validated)

    # Q1: MFC 합침 + ATTRIBUTE_CONFLICT 분류 (v3 NEW-M-1 정합)
    merged_rows, conflict_rows = _merge_duplicate_mfc(parsed_rows)
    rejected_rows.extend(conflict_rows)

    return merged_rows, rejected_rows
